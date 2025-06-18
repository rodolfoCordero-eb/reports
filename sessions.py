import boto3
import pandas as pd
from botocore.exceptions import ClientError
import os
from openpyxl import load_workbook

class Session:
    def __init__(self):
        self.ROLE_NAME="OrganizationAccountAccessRole"
        self.org_client = boto3.client("organizations")
        self.sts_client = boto3.client("sts")
        self.set_caller_identity()
        self.list_accounts()
        self.file_paths="./excel_files"
        self.draw_path="./images"
        pass

    def assume_role(self,account_id, role_name):
        try:
            response = self.sts_client.assume_role(
                RoleArn=f"arn:aws:iam::{account_id}:role/{role_name}",
                RoleSessionName="OrgAuditSession"
            )
            creds = response["Credentials"]
            return boto3.Session(
                aws_access_key_id=creds["AccessKeyId"],
                aws_secret_access_key=creds["SecretAccessKey"],
                aws_session_token=creds["SessionToken"],
            )
        except ClientError as e:
            print(f"[{account_id}] ❌ Failed to assume role: {e}")
            return None

    def list_accounts(self):
        accounts = []
        paginator = self.org_client.get_paginator("list_accounts")
        for page in paginator.paginate():
            accounts.extend(page["Accounts"])
        self.accounts = [acc for acc in accounts if acc["Status"] == "ACTIVE"]

    def set_caller_identity(self):
        self.identity = self.sts_client.get_caller_identity()
        userId=self.identity["UserId"]
        account = self.identity["Account"]
        print(f"\n👤 UserId: {userId}")
        print(f"\n🧾 AccountId: {account}")
    
    def execute(self,strategy):
        if "draw" in strategy.__class__.__name__:
            print("Strategy Draw")
            self.execute_draw(strategy)
        if "action" in strategy.__class__.__name__:
            print("Strategy Action")
            self.execute_excel(strategy)

    def execute_excel(self,strategy):
        print("\n\n----------------------------------")
        print(f"👓 {strategy.name()}")
        file_name = f"{self.file_paths}/{strategy.name()}.xlsx"
        if not os.path.exists(file_name):
            print("📗 Creating book")
            df = pd.DataFrame({"Module":[strategy.name()]})
            with pd.ExcelWriter(file_name,engine='openpyxl') as writer:
                df.to_excel(writer,sheet_name="Index",index=False)

        for acc in self.accounts:
            acc_id = acc["Id"]
            acc_name = acc["Name"]
            print(f"🔍 Scanning account {acc_id} ({acc_name})...!!!!!!")
            session  =self.assume_role(acc_id,self.ROLE_NAME)
            if (session):
                result = strategy.run(session,acc_id, acc_name)
                if not result.empty:
                    book = load_workbook(file_name)
                    if acc_name in book.sheetnames:
                        print(f"🗡️ Remove book {acc_name}")
                        std = book[acc_name]
                        book.remove(std)
                        book.save(file_name)
                    print(f"\n📁 Writing File {strategy.name()} with {acc_name} content")
                    with pd.ExcelWriter(file_name,engine='openpyxl', mode = 'a') as writer:
                        result.to_excel(writer,sheet_name=acc_name,index=False)
        print("----------------------------------\n\n")

    def execute_draw(self,strategy):
        print("\n\n----------------------------------")
        print(f"👓 {strategy.name()} Draw")

        for acc in self.accounts:
            acc_id = acc["Id"]
            acc_name = acc["Name"]
            print(f"🔍 Scanning account {acc_id} ({acc_name})...!!")
            session  =self.assume_role(acc_id,self.ROLE_NAME)
            if (session):
                strategy.run(session,acc_id, acc_name,self.draw_path)
        print("----------------------------------\n\n")
