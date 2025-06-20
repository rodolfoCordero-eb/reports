import boto3
import pandas as pd
from botocore.exceptions import ClientError
from openpyxl import load_workbook
import os

class Single:
    def __init__(self,session=None,name=None):
        self.account_name=name
        if not session:
            session= boto3.Session()
        self.session=session
        self.set_caller_identity()
        self.file_paths="./excel_files"
        self.draw_path="./images"
        pass

    def set_caller_identity(self):
        org = self.session.client('organizations')
        sts_client = self.session.client("sts")
        self.identity = sts_client.get_caller_identity()

        self.userId=self.identity["UserId"]
        self.account_id = self.identity["Account"]
        if not self.account_name:
            self.account_name = org.describe_account(AccountId=self.account_id).get('Account').get('Name')

        
        print(f"🔍 Scanning account {self.account_id}-({self.account_name}) {self.account_name}...!!")
        return self.identity
    
    def execute(self,strategy,region=None):
        if "draw" in strategy.__class__.__name__:
            print("Strategy Draw")
            self.write_draw(strategy,region)
        if "action" in strategy.__class__.__name__:
            print("Strategy Action")
            self.write_excel(strategy)

    def write_draw(self,strategy,region=None):
        print("\n\n----------------------------------")
        print(f"👓 {strategy.name()} Draw")
        strategy.run(self.session,self.account_id, self.account_name,self.draw_path,region)
        print("----------------------------------\n\n")

    def write_excel(self,strategy):
        file_name = f"{self.file_paths}/{strategy.name()}.xlsx"
        print("\n\n----------------------------------")
        print(f"👓 {strategy.name()}")
        if not os.path.exists(file_name):
            print("📗 Creating book")
            df = pd.DataFrame({"Module":[strategy.name()]})
            with pd.ExcelWriter(file_name,engine='openpyxl') as writer:
                df.to_excel(writer,sheet_name="Index",index=False)          
        result = strategy.run(self.session,self.account_id, self.account_name)
        if not result.empty:
            book = load_workbook(file_name)
            if self.account_name in book.sheetnames:
                print(f"🗡️ Remove book {self.account_name}")
                std = book[self.account_name]
                book.remove(std)
                book.save(file_name)
            print(f"\n📁 Writing File {strategy.name()} with {self.account_name} content")
            with pd.ExcelWriter(file_name,engine='openpyxl', mode = 'a') as writer:
                result.to_excel(writer,sheet_name=self.account_name,index=False)
        print("----------------------------------\n\n")
